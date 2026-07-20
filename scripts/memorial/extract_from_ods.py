# -*- coding: utf-8 -*-
"""ODS maître « Morts pour la France » -> 5 Excel propres (4 colonnes, tri alphabétique).

Usage :
  python scripts/memorial/extract_from_ods.py <fichier.ods> [--sortie data-memorial]

Filtrage (validé le 18/07/2026) :
  - Feuilles à flag numérique (col A)      : ligne retenue ssi flag == 1.
  - Feuilles Opex sans flag exploitable    : ligne retenue si un Nom est présent
    et que le flag n'est pas un nombre != 1 (mode « souple »).
  - « Feuille25 » : exclue (décision utilisateur).
  - « Conflit _Riff… », « Afrique Occidentale Française », « Cameroun » : flag
    numérique sans aucun 1 -> naturellement vides après filtrage.
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).parent))
from common import (COLONNES_PROPRES, FICHIERS_CATEGORIE, parse_date_deces,
                    texte, trie_personnes)

# ── Configuration par feuille : catégorie + mode de flag ──
# en-tête localisé par nom de colonne ; « index_fixes » pour les feuilles sans en-tête.
FEUILLES = {
    "Fiches_vérifiées_1914_1918": {"cat": "1GM", "flag": "strict1"},
    "Liste_1939_1945": {"cat": "2GM", "flag": "strict1"},
    # "Feuille25" : exclue (décision utilisateur du 18/07/2026)
    "indochine": {"cat": "Indochine", "flag": "strict1"},
    "Algérie": {"cat": "Algérie", "flag": "strict1"},
    # « conflit » : théâtre affiché sous chaque nom de l'onglet Opex
    # (colonne « Conflit » du fichier propre) — libellé issu de la feuille source.
    "Conflit_Levant": {
        "cat": "Opex", "flag": "strict1", "conflit": "Levant",
        "index_fixes": {"flag": 0, "nom": 5, "prenom": 6, "grade": 7, "an": 17, "date": 18},
    },
    "Conflit _Riff_1920_1930_1939": {"cat": "Opex", "flag": "strict1", "conflit": "Guerre du Rif"},
    "Afrique Occidentale Française": {"cat": "Opex", "flag": "strict1",
                                      "conflit": "Afrique-Occidentale française"},
    "Cameroun": {"cat": "Opex", "flag": "strict1", "conflit": "Cameroun"},
    "Yougoslavie": {"cat": "Opex", "flag": "strict1", "conflit": "Ex-Yougoslavie"},
    "Levant": {"cat": "Opex", "flag": "souple", "conflit": "Levant"},
    "SOUDAN": {"cat": "Opex", "flag": "souple", "conflit": "Soudan"},
    "Tunisie": {"cat": "Opex", "flag": "souple", "conflit": "Tunisie"},
    "Corée_ONU": {"cat": "Opex", "flag": "souple", "conflit": "Corée (ONU)"},
    "Tchad": {"cat": "Opex", "flag": "souple", "conflit": "Tchad"},
    "Afganistan": {"cat": "Opex", "flag": "souple", "conflit": "Afghanistan"},
}

# Alias d'en-têtes (clé normalisée : minuscules, sans espaces).
ALIAS = {
    "nom": "nom",
    "prenom": "prenom", "prenoms": "prenom", "prénom": "prenom", "prénoms": "prenom",
    "grade": "grade",
    "andc": "an", "année": "an", "annee": "an",
    "datedc(mpf)": "date", "datedc": "date",
}


def _norme(h: str) -> str:
    return re.sub(r"\s+", "", texte(h)).lower()


def localise_colonnes(df: pd.DataFrame) -> tuple[int, dict[str, int]]:
    """Trouve la ligne d'en-tête (parmi les 3 premières) et mappe les colonnes."""
    for i in range(min(3, len(df))):
        cellules = [_norme(v) for v in df.iloc[i].tolist()]
        if "nom" in cellules:
            colonnes: dict[str, int] = {"flag": 0}
            for idx, cell in enumerate(cellules):
                cle = ALIAS.get(cell)
                if cle and cle not in colonnes:
                    colonnes[cle] = idx
            return i, colonnes
    raise ValueError("ligne d'en-tête introuvable (aucune cellule « Nom »)")


def flag_retenu(mode: str, flag: str, nom: str) -> bool:
    if flag == "Id1" or not nom or nom.lower() in ("nom",):
        return False
    if mode == "strict1":
        return flag == "1"
    # mode « souple » : flag vide ou 1 -> oui ; nombre != 1 -> non.
    if re.fullmatch(r"\d+", flag):
        return flag == "1"
    return True  # vide ou texte non numérique, avec un nom présent


def extrait(chemin_ods: Path) -> dict[str, list[dict]]:
    feuilles = pd.read_excel(chemin_ods, sheet_name=None, engine="odf",
                             header=None, dtype=object)
    categories: dict[str, list[dict]] = {c: [] for c in FICHIERS_CATEGORIE}
    for nom_feuille, cfg in FEUILLES.items():
        if nom_feuille not in feuilles:
            print(f"  AVERTISSEMENT : feuille absente du classeur : {nom_feuille!r}")
            continue
        df = feuilles[nom_feuille]
        if "index_fixes" in cfg:
            debut, cols = 0, cfg["index_fixes"]
        else:
            ligne_entete, cols = localise_colonnes(df)
            debut = ligne_entete + 1
        retenues = 0
        for i in range(debut, len(df)):
            ligne = df.iloc[i]

            def val(cle: str) -> str:
                idx = cols.get(cle)
                return texte(ligne.iloc[idx]) if idx is not None and idx < len(ligne) else ""

            nom = val("nom")
            if not flag_retenu(cfg["flag"], val("flag"), nom):
                continue
            idx_date = cols.get("date")
            brut_date = ligne.iloc[idx_date] if idx_date is not None and idx_date < len(ligne) else None
            categories[cfg["cat"]].append({
                "nom": nom.upper(),
                "prenom": val("prenom"),
                "grade": val("grade"),
                "date": parse_date_deces(brut_date, val("an")),
                "conflit": cfg.get("conflit", ""),
            })
            retenues += 1
        print(f"  {nom_feuille!r} -> {cfg['cat']} : {retenues} retenue(s)")
    return categories


def ecrit_xlsx(personnes: list[dict], chemin: Path, categorie: str) -> None:
    """Fichier Excel « propre » : 1 feuille, 4 colonnes (+ « Conflit » si renseigné,
    cas de l'Opex), sans formule."""
    avec_conflit = any(p.get("conflit") for p in personnes)
    colonnes = COLONNES_PROPRES + (["Conflit"] if avec_conflit else [])
    wb = Workbook()
    ws = wb.active
    ws.title = categorie
    ws.append(colonnes)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D3151")
    for p in personnes:
        ligne = [p["nom"], p["prenom"], p["date"], p["grade"]]
        if avec_conflit:
            ligne.append(p.get("conflit", ""))
        ws.append(ligne)
    for col, largeur in zip("ABCDE", (30, 32, 16, 28, 24)):
        ws.column_dimensions[col].width = largeur
    for ligne in ws.iter_rows(min_row=2):
        for cell in ligne:
            cell.font = Font(name="Arial")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{'E' if avec_conflit else 'D'}{ws.max_row}"
    chemin.parent.mkdir(parents=True, exist_ok=True)
    wb.save(chemin)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("ods", help="fichier ODS maître (feuilles par conflit)")
    ap.add_argument("--sortie", default=str(Path(__file__).parents[2] / "data-memorial"))
    args = ap.parse_args()

    print(f"Lecture de {args.ods}")
    categories = extrait(Path(args.ods))

    sortie = Path(args.sortie)
    for cat, nom_fichier in FICHIERS_CATEGORIE.items():
        personnes = trie_personnes(categories[cat])
        chemin = sortie / f"{nom_fichier}.xlsx"
        ecrit_xlsx(personnes, chemin, cat)
        print(f"  {cat} : {len(personnes)} personnes -> {chemin}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
